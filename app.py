"""
気象レポートPDF 降水量ハイライター
複数PDF・PDF/JPG出力・週間天気予報PDF生成・1ページ合体対応版
"""

import streamlit as st
import tempfile
import os
import io
import zipfile
import subprocess
from itertools import groupby

import pdfplumber
import pypdfium2 as pdfium
import pypdfium2.raw as pdfium_c
from PIL import Image, ImageChops
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from pypdf import PdfWriter, PdfReader

st.set_page_config(
    page_title="降水量ハイライター",
    page_icon="🌧️",
    layout="centered"
)

st.markdown("""
<style>
    .stButton > button {
        width: 100%;
        background: linear-gradient(135deg, #00B0F0, #0097d6);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 14px;
        font-size: 16px;
        font-weight: 600;
    }
    .stButton > button:hover { opacity: 0.9; }
    .result-box {
        background: #e8f9f4;
        border: 1px solid #c3f0e4;
        border-radius: 12px;
        padding: 20px;
        text-align: center;
        margin: 16px 0;
    }
    .file-row {
        background: #f4f9fd;
        border: 1px solid #cde4f5;
        border-radius: 8px;
        padding: 10px 14px;
        margin: 6px 0;
        font-size: 13px;
    }
    .section-title {
        font-size: 15px;
        font-weight: 700;
        color: #0a1628;
        margin: 24px 0 8px;
        padding-bottom: 6px;
        border-bottom: 2px solid #cde4f5;
    }
</style>
""", unsafe_allow_html=True)


# ============================================================
# PDF処理ロジック
# ============================================================

def extract_rainfall_groups(pdf_path):
    page_groups = []
    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            words = page.extract_words()
            mmh_tops = [w['top'] for w in words if w['text'] == 'mm/h']
            if not mmh_tops:
                continue
            tset = {"vertical_strategy": "lines", "horizontal_strategy": "lines"}
            tables = page.find_tables(tset)
            if not tables:
                continue
            all_cells = tables[0].cells
            rainfall_boxes = []
            for w in words:
                in_row = any(abs(w['top'] - mt) <= 8 for mt in mmh_tops)
                if not in_row:
                    continue
                try:
                    val = float(w['text'])
                    if val <= 0:
                        continue
                    cx = (w['x0'] + w['x1']) / 2
                    cy = (w['top'] + w['bottom']) / 2
                    for cell in all_cells:
                        if cell[0] <= cx <= cell[2] and cell[1] <= cy <= cell[3]:
                            rainfall_boxes.append({
                                'value': val,
                                'x0': cell[0], 'top': cell[1],
                                'x1': cell[2], 'bottom': cell[3],
                                'row_top': round(cell[1], 1)
                            })
                            break
                except ValueError:
                    pass
            groups = []
            boxes_sorted = sorted(rainfall_boxes, key=lambda c: (c['row_top'], c['x0']))
            for _, row_boxes in groupby(boxes_sorted, key=lambda c: c['row_top']):
                row_boxes = list(row_boxes)
                current = [row_boxes[0]]
                for box in row_boxes[1:]:
                    if box['x0'] - current[-1]['x1'] <= 25:
                        current.append(box)
                    else:
                        groups.append(current)
                        current = [box]
                groups.append(current)
            page_groups.append({
                'page_num': page_num,
                'page_height': page.height,
                'groups': groups
            })
    return page_groups


def draw_highlights(input_path, output_path, page_groups):
    doc = pdfium.PdfDocument(input_path)
    SKY_R, SKY_G, SKY_B = 0, 176, 240
    for pg in page_groups:
        page = doc[pg['page_num']]
        pdf_h = page.get_height()
        pl_h = pg['page_height']
        for group in pg['groups']:
            gx0 = min(c['x0'] for c in group)
            gx1 = max(c['x1'] for c in group)
            gtop = min(c['top'] for c in group)
            gbottom = max(c['bottom'] for c in group)
            pdf_x0 = gx0
            pdf_y0 = pdf_h - gbottom * (pdf_h / pl_h)
            pdf_y1 = pdf_h - gtop * (pdf_h / pl_h)
            rect = pdfium_c.FPDFPageObj_CreateNewRect(
                pdf_x0, pdf_y0, gx1 - gx0, pdf_y1 - pdf_y0
            )
            pdfium_c.FPDFPageObj_SetFillColor(rect, SKY_R, SKY_G, SKY_B, 40)
            pdfium_c.FPDFPageObj_SetStrokeColor(rect, SKY_R, SKY_G, SKY_B, 255)
            pdfium_c.FPDFPageObj_SetStrokeWidth(rect, 1.5)
            pdfium_c.FPDFPath_SetDrawMode(rect, 1, 1)
            pdfium_c.FPDFPageObj_SetBlendMode(rect, b"Multiply")
            pdfium_c.FPDFPage_InsertObject(page.raw, rect)
        pdfium_c.FPDFPage_GenerateContent(page.raw)
    doc.save(output_path)


def pdf_to_jpg_bytes(pdf_path, dpi=150):
    doc = pdfium.PdfDocument(pdf_path)
    scale = dpi / 72
    results = []
    for page_num in range(len(doc)):
        page = doc[page_num]
        bitmap = page.render(scale=scale, rotation=0)
        pil_image = bitmap.to_pil()
        buf = io.BytesIO()
        pil_image.save(buf, format="JPEG", quality=90)
        results.append((page_num, buf.getvalue()))
    return results


def trim_whitespace(img, margin=5):
    bg = Image.new("RGB", img.size, (255, 255, 255))
    diff = ImageChops.difference(img, bg)
    bbox = diff.getbbox()
    if bbox is None:
        return img
    left   = max(0, bbox[0] - margin)
    top    = max(0, bbox[1] - margin)
    right  = min(img.width,  bbox[2] + margin)
    bottom = min(img.height, bbox[3] + margin)
    return img.crop((left, top, right, bottom))


def crop_forecast_image(pdf_path, dpi=200):
    doc = pdfium.PdfDocument(pdf_path)
    page = doc[0]
    scale = dpi / 72
    bitmap = page.render(scale=scale)
    img = bitmap.to_pil()
    px_top = int(212 * scale)
    px_bottom = int(780 * scale)
    cropped = img.crop((0, px_top, img.width, px_bottom))
    return trim_whitespace(cropped, margin=5)


def embed_to_excel(highlighted_pdf_path, template_bytes, tmpdir):
    crop_img = crop_forecast_image(highlighted_pdf_path, dpi=200)
    img_path = os.path.join(tmpdir, "forecast_crop.png")
    crop_img.save(img_path, "PNG")

    template_path = os.path.join(tmpdir, "template.xlsm")
    with open(template_path, "wb") as f:
        f.write(template_bytes)

    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb['出力']

    ws._images = [
        im for im in ws._images
        if hasattr(im.anchor, '_from') and im.anchor._from.row < 23
    ]

    TARGET_WIDTH_PX = 1280
    orig_w, orig_h = crop_img.size
    ratio = TARGET_WIDTH_PX / orig_w
    target_height_px = int(orig_h * ratio)

    xl_img = XLImage(img_path)
    xl_img.width = TARGET_WIDTH_PX
    xl_img.height = target_height_px
    xl_img.anchor = "A24"
    ws.add_image(xl_img)

    xlsm_path = os.path.join(tmpdir, "週間天気予報_完成.xlsm")
    wb.save(xlsm_path)

    with open(xlsm_path, "rb") as f:
        xlsm_bytes = f.read()

    return xlsm_path, xlsm_bytes


def excel_to_pdf(xlsm_path, tmpdir):
    try:
        wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)
        for sheet_name in wb.sheetnames:
            if sheet_name != '出力':
                wb[sheet_name].sheet_state = 'hidden'

        xlsx_path = os.path.join(tmpdir, "for_pdf.xlsx")
        wb.save(xlsx_path)

        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", tmpdir, xlsx_path],
            capture_output=True, text=True, timeout=60
        )

        pdf_path = os.path.join(tmpdir, "for_pdf.pdf")
        if os.path.exists(pdf_path):
            with open(pdf_path, "rb") as f:
                return f.read()
        return None
    except Exception:
        return None


def combine_as_one_page(weekly_pdf_bytes, report_pdf_bytes, dpi=150):
    """
    週間天気予報PDFの下に1時間予報の切り抜きを貼り付けて
    1ページのPDFを生成して返す
    """
    scale = dpi / 72

    # 週間天気予報PDFをレンダリング（1ページ目）
    weekly_doc = pdfium.PdfDocument(io.BytesIO(weekly_pdf_bytes))
    weekly_page = weekly_doc[0]
    weekly_img = weekly_page.render(scale=scale).to_pil()

    # 気象レポートPDFにハイライト処理して1時間予報を切り抜き
    with tempfile.TemporaryDirectory() as tmpdir:
        input_path = os.path.join(tmpdir, "report.pdf")
        highlighted_path = os.path.join(tmpdir, "highlighted.pdf")

        with open(input_path, "wb") as f:
            f.write(report_pdf_bytes)

        page_groups = extract_rainfall_groups(input_path)
        if page_groups:
            draw_highlights(input_path, highlighted_path, page_groups)
        else:
            highlighted_path = input_path

        crop_img = crop_forecast_image(highlighted_path, dpi=dpi)

    # 切り抜き画像を週間天気予報の幅に合わせてリサイズ
    target_w = weekly_img.width
    ratio = target_w / crop_img.width
    crop_resized = crop_img.resize(
        (target_w, int(crop_img.height * ratio)), Image.LANCZOS
    )

    # 縦に結合（週間天気予報の下に1時間予報を配置）
    combined = Image.new(
        "RGB",
        (target_w, weekly_img.height + crop_resized.height),
        (255, 255, 255)
    )
    combined.paste(weekly_img, (0, 0))
    combined.paste(crop_resized, (0, weekly_img.height))

    # 1ページのPDFとして保存
    buf = io.BytesIO()
    combined.save(buf, format="PDF", resolution=dpi)
    return buf.getvalue()


def process_single_pdf(uploaded_file, tmpdir, output_format, dpi=150):
    input_path = os.path.join(tmpdir, "input.pdf")
    highlighted_path = os.path.join(tmpdir, "highlighted.pdf")
    base_name = os.path.splitext(uploaded_file.name)[0]

    with open(input_path, "wb") as f:
        f.write(uploaded_file.getbuffer())

    page_groups = extract_rainfall_groups(input_path)
    total_cells = sum(len(g) for pg in page_groups for g in pg['groups'])
    total_groups = sum(len(pg['groups']) for pg in page_groups)

    if total_cells == 0:
        return None, 0, 0, None

    draw_highlights(input_path, highlighted_path, page_groups)

    outputs = []
    if output_format == "PDF":
        with open(highlighted_path, "rb") as f:
            outputs.append({
                "name": f"{base_name}_ハイライト済.pdf",
                "bytes": f.read(),
                "mime": "application/pdf"
            })
    else:
        pages = pdf_to_jpg_bytes(highlighted_path, dpi=dpi)
        if len(pages) == 1:
            _, jpg_bytes = pages[0]
            outputs.append({
                "name": f"{base_name}_ハイライト済.jpg",
                "bytes": jpg_bytes,
                "mime": "image/jpeg"
            })
        else:
            for page_num, jpg_bytes in pages:
                outputs.append({
                    "name": f"{base_name}_ハイライト済_p{page_num + 1}.jpg",
                    "bytes": jpg_bytes,
                    "mime": "image/jpeg"
                })

    return outputs, total_cells, total_groups, highlighted_path


# ============================================================
# 画面
# ============================================================

st.title("🌧️ 降水量ハイライター")
st.caption("気象レポートのPDFをアップロードすると、降水量が0より大きいセルを水色の枠と薄い水色でハイライトします。")

st.divider()

# モード選択
st.markdown('<div class="section-title">処理モードを選択</div>', unsafe_allow_html=True)
mode = st.radio(
    label="モード",
    options=["📄 ハイライトのみ", "📊 週間天気予報PDFも作成", "📎 1ページに合体させる"],
    label_visibility="collapsed",
    horizontal=True
)

st.divider()

# ============================================================
# モード：1ページに合体させる
# ============================================================
if mode == "📎 1ページに合体させる":
    st.markdown('<div class="section-title">① 週間天気予報PDF（出力シートをPDF化したもの）</div>', unsafe_allow_html=True)
    weekly_pdf_file = st.file_uploader(
        "週間天気予報のPDFを選択",
        type=["pdf"],
        key="weekly_pdf_upload"
    )

    st.markdown('<div class="section-title">② 気象レポートPDF</div>', unsafe_allow_html=True)
    st.caption("ハイライト処理と1時間予報の切り抜きを自動で行います")
    report_pdf_file = st.file_uploader(
        "気象レポートのPDFを選択",
        type=["pdf"],
        key="report_pdf_upload"
    )

    st.markdown("")

    if weekly_pdf_file and report_pdf_file:
        st.success(f"✅ {weekly_pdf_file.name}  ＋  {report_pdf_file.name}")

        if st.button("📎 1ページに合体してダウンロード"):
            with st.spinner("処理中... ハイライト処理と合体を行っています"):
                try:
                    result_bytes = combine_as_one_page(
                        weekly_pdf_file.getbuffer(),
                        report_pdf_file.getbuffer(),
                        dpi=150
                    )

                    st.markdown("""
                    <div class="result-box">
                        <h3 style="color:#00c896; margin:0 0 8px">✓ 合体完了！</h3>
                        <p style="color:#6b8aad; margin:0">週間天気予報の下に1時間予報を配置した1ページPDFが完成しました</p>
                    </div>
                    """, unsafe_allow_html=True)

                    st.download_button(
                        label="📥 完成PDFをダウンロード",
                        data=result_bytes,
                        file_name="週間天気予報_完成.pdf",
                        mime="application/pdf"
                    )
                except Exception as e:
                    st.error(f"エラーが発生しました: {e}")

# ============================================================
# モード：ハイライトのみ / 週間天気予報PDFも作成
# ============================================================
else:
    st.markdown('<div class="section-title">① 気象レポートPDFをアップロード</div>', unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        "気象レポートのPDFを選択してください（複数選択可）",
        type=["pdf"],
        accept_multiple_files=True,
        help="Ctrl（Mac: Command）を押しながらクリックで複数選択できます"
    )

    template_file = None
    if mode == "📊 週間天気予報PDFも作成":
        st.markdown('<div class="section-title">② 週間天気予報テンプレートをアップロード</div>', unsafe_allow_html=True)
        template_file = st.file_uploader(
            "週間天気予報.xlsm を選択してください",
            type=["xlsm", "xlsx"],
            help="毎回同じテンプレートファイルを使用します"
        )

    if uploaded_files:
        st.markdown(f"**{len(uploaded_files)} 件のファイルが選択されています**")
        for f in uploaded_files:
            st.markdown(f"""
            <div class="file-row">
                📄 {f.name}&nbsp;&nbsp;<span style="color:#6b8aad">{f.size / 1024:.0f} KB</span>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("")

        if mode == "📄 ハイライトのみ":
            st.markdown("**出力形式を選択してください**")
            output_format = st.radio(
                label="出力形式",
                options=["PDF", "JPG"],
                horizontal=True,
                label_visibility="collapsed"
            )
            dpi = st.select_slider(
                "画質（DPI）",
                options=[72, 96, 150, 200, 300],
                value=150,
                help="JPG選択時のみ有効"
            ) if output_format == "JPG" else 150
        else:
            output_format = "PDF"
            dpi = 150

        st.markdown("")

        btn_label = "⚡ ハイライト処理を実行" if mode == "📄 ハイライトのみ" else "⚡ ハイライト＋週間天気予報を作成"
        run = st.button(btn_label)

        if run:
            if mode == "📊 週間天気予報PDFも作成" and not template_file:
                st.error("⚠️ 週間天気予報のテンプレートファイルをアップロードしてください。")
                st.stop()
            if mode == "📊 週間天気予報PDFも作成" and len(uploaded_files) > 1:
                st.error("⚠️ 週間天気予報作成モードは気象レポートを1ファイルずつ処理してください。")
                st.stop()

            all_outputs = []
            weekly_excel_bytes = None
            weekly_pdf_bytes = None
            errors = []

            progress_bar = st.progress(0, text="処理中...")

            with tempfile.TemporaryDirectory() as tmpdir:
                for i, uploaded_file in enumerate(uploaded_files):
                    progress_bar.progress(
                        i / len(uploaded_files),
                        text=f"処理中 ({i+1}/{len(uploaded_files)}): {uploaded_file.name}"
                    )
                    try:
                        outputs, total_cells, total_groups, highlighted_path = process_single_pdf(
                            uploaded_file, tmpdir, output_format, dpi
                        )
                        if outputs is None:
                            errors.append(f"{uploaded_file.name}：降水量データが見つかりませんでした")
                            continue

                        for out in outputs:
                            all_outputs.append({**out, "cells": total_cells, "groups": total_groups})

                        if mode == "📊 週間天気予報PDFも作成" and highlighted_path:
                            progress_bar.progress(0.6, text="Excelテンプレートに貼り付け中...")
                            xlsm_path, weekly_excel_bytes = embed_to_excel(
                                highlighted_path, template_file.getbuffer(), tmpdir
                            )
                            progress_bar.progress(0.85, text="PDFに変換中...")
                            weekly_pdf_bytes = excel_to_pdf(xlsm_path, tmpdir)

                    except Exception as e:
                        errors.append(f"{uploaded_file.name}：エラー ({e})")

                progress_bar.progress(1.0, text="完了！")

                for err in errors:
                    st.error(f"⚠️ {err}")

                if all_outputs or weekly_excel_bytes:
                    total_files = len(uploaded_files) - len(errors)
                    st.markdown(f"""
                    <div class="result-box">
                        <h3 style="color:#00c896; margin:0 0 8px">✓ {total_files} 件の処理が完了しました！</h3>
                        <p style="color:#6b8aad; margin:0">合計 {sum(o['cells'] for o in all_outputs)} セルをハイライト</p>
                    </div>
                    """, unsafe_allow_html=True)

                    if all_outputs:
                        st.markdown("**📄 ハイライト済みファイル**")
                        if len(all_outputs) == 1:
                            out = all_outputs[0]
                            st.download_button(
                                label=f"📥 {out['name']} をダウンロード",
                                data=out['bytes'],
                                file_name=out['name'],
                                mime=out['mime']
                            )
                        else:
                            zip_buf = io.BytesIO()
                            with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                                for out in all_outputs:
                                    zf.writestr(out['name'], out['bytes'])
                            zip_buf.seek(0)
                            st.download_button(
                                label=f"📦 {len(all_outputs)} 件をZIPでまとめてダウンロード",
                                data=zip_buf.getvalue(),
                                file_name="ハイライト済み_一括.zip",
                                mime="application/zip"
                            )
                            with st.expander("📄 個別にダウンロードする"):
                                for out in all_outputs:
                                    col1, col2 = st.columns([3, 1])
                                    with col1:
                                        st.caption(out['name'])
                                    with col2:
                                        st.download_button(
                                            label="DL",
                                            data=out['bytes'],
                                            file_name=out['name'],
                                            mime=out['mime'],
                                            key=out['name']
                                        )

                    if weekly_excel_bytes or weekly_pdf_bytes:
                        st.markdown("**📊 週間天気予報**")
                        col1, col2 = st.columns(2)
                        with col1:
                            if weekly_pdf_bytes:
                                st.download_button(
                                    label="📥 PDF でダウンロード",
                                    data=weekly_pdf_bytes,
                                    file_name="週間天気予報_完成.pdf",
                                    mime="application/pdf",
                                    key="weekly_pdf"
                                )
                            else:
                                st.warning("PDF変換に失敗しました。Excelファイルをご利用ください。")
                        with col2:
                            if weekly_excel_bytes:
                                st.download_button(
                                    label="📥 Excel でダウンロード",
                                    data=weekly_excel_bytes,
                                    file_name="週間天気予報_完成.xlsm",
                                    mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                                    key="weekly_excel"
                                )

st.divider()
st.caption("使い方：モード選択 -> PDFをアップロード -> 実行 -> ダ
